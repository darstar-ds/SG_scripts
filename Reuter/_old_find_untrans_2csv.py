import pandas as pd
import os
import openpyxl
import datetime

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Langbeschreibung",
    11: "Ergaenzung-Bullets",
    12: "Weitere-Anmerkungen",
    13: "Achtung",
    14: "Variante",
    15: "Hinweis",
    16: "Typ",
    17: "Abmessungen",
    18: "Leistung-Leuchtmittel"
    }

CURR_PATH = "d:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-07-26\\"
EXPORTS4TRANS = CURR_PATH + "exports4trans\\"
EXPORTED_COLUMNS = CURR_PATH + "4trans_sheets\\"
TRANSCOLUMNS = list(XLS_COLUMNS.keys())


def find_untrans(files4trans_path, columns):
    '''
    Provide the path to the directory
    with xlsx exports for trans.
    Source cells that require translation
    will be exported
    '''
    prev_time = datetime.datetime.now()
    dir_list = os.listdir(files4trans_path) 
    if os.path.exists(files4trans_path):
        print(f"Folder {files4trans_path} exist.")
    else:
        print(f"Folder {files4trans_path} does not exist.")
    print(f"dir_list= {dir_list}")
    
    for file in dir_list: # get all untranslated entries and build separate workbook
        file4trans_path = files4trans_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        
        for tab in wb.sheetnames: #for each worksheet in a workbook
            ws = wb[tab]
            row_limit = ws.max_row
            print(f"Number of rows: {row_limit} \nCurrent sheet: {ws} \nCurrent tab: {tab}")
            
            # verify the names/number of columns in the spreadsheet
            list_with_values=[]
            for cell in ws[1]:
                list_with_values.append(cell.value)
            print(f"Columns in {tab} tab: {list_with_values}")
            print(f"Columns no in {tab} tab: {len(list_with_values)}")
    
            for transcolumn in columns:
                strings4trans = pd.DataFrame(columns=["DE", "PL"])
                csv_path = EXPORTED_COLUMNS + file[:-5] + "-" + tab + "-" + "COL" + "-" + str(XLS_COLUMNS[transcolumn]) + ".csv"
                strings4trans.to_csv(csv_path, sep='\t', header=True, index=False)
                curr_row = 2
                while curr_row < row_limit:
                    # print(f"Sprache type: {str(ws.cell(row=curr_row, column=2))}")
                    # print(f"Sprache value: {str(ws.cell(row=curr_row, column=2).value)}")
                    if ws.cell(row=curr_row, column=2).value == "deu" and \
                        ws.cell(row=curr_row, column=transcolumn).value is not None:
                        source_datenart = ws.cell(row=curr_row, column=1).value
                        source_sprache = ws.cell(row=curr_row, column=2).value
                        source_produktno = ws.cell(row=curr_row, column=3).value
                        source_navisionid = ws.cell(row=curr_row, column=4).value
                        source_text = ws.cell(row=curr_row, column=transcolumn).value
                        # print(f"Source_text type: {type(source_text)}, source_text: {source_text}")
                        
                        is_target_found = False
                        while is_target_found == False:
                            for row in ws.iter_rows(min_row=curr_row, max_row=row_limit):
                                # print(row)
                                # target_datenart = ws.cell(row=row, column=1).value
                                target_datenart = row[0].value
                                # print(f"target_datenart: {target_datenart}")
                                target_sprache = row[1].value
                                target_produktno = row[2].value
                                target_navisionid = row[3].value
                                target_text = row[transcolumn].value
                            
                                if target_datenart == source_datenart and \
                                    target_sprache == "pol" and \
                                    target_produktno == source_produktno and \
                                    target_navisionid == source_navisionid and \
                                    target_text is None:
                                        target_text = ""
                                        # print(f"Target_text type: {type(target_text)}, target_text: {target_text}")
                                        # print(row)                                    
                                        # new_row = {"DE": source_text, "PL": target_text}
                                        data = {"DE": [source_text], "PL": [target_text]}
                                        # data=[source_text, target_text], columns=["DE", "PL"]
                                        new_row = pd.DataFrame(data)
                                        # strings4trans = strings4trans.append(new_row, ignore_index=True)
                                        new_row.to_csv(csv_path, mode="a", sep='\t', index=False, header=False)
                                        # print(strings4trans)
                                        is_target_found=True
                    else:
                        pass        
                    curr_row += 1
                    if curr_row % 200 == 0:
                        curr_time = datetime.datetime.now()
                        time_delta = curr_time - prev_time
                        print(f"Column: {XLS_COLUMNS[transcolumn]}, Progress: {curr_row}/{row_limit}, Time: {curr_time}, Delta: {time_delta}")
                        prev_time = curr_time
                
                csv_path_df = pd.read_csv(csv_path, sep="\t")
                strings4trans = pd.concat([strings4trans, csv_path_df], ignore_index=True)
                print(f"Tab: {tab}, Column: {XLS_COLUMNS[transcolumn]}, initial dataframe head: {strings4trans.head()}")
                print(f"Initial dataframe shape: {strings4trans.shape}")

        
                # extract unique entries only
                strings4trans_unique = strings4trans.drop_duplicates(keep="first")
                print(f"Tab: {tab}, Column: {XLS_COLUMNS[transcolumn]}, unique dataframe head: {strings4trans_unique.head()}")
                print(f"Unique dataframe shape: {strings4trans_unique.shape}")

                # save the dataframe to excel
                if strings4trans_unique.shape[0] == 0:
                    pass
                else:
                    file_path = EXPORTED_COLUMNS + file[:-5] + "-" + tab + "-" + "COL" + "-" + str(XLS_COLUMNS[transcolumn]) + "_unique.xlsx"
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        strings4trans_unique.to_excel(writer, sheet_name=str(XLS_COLUMNS[transcolumn]), index=False)
                

find_untrans(EXPORTS4TRANS, TRANSCOLUMNS)