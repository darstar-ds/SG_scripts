import pandas as pd
import os
import openpyxl

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Langbeschreibung",
    11: "Ergaenzung-Bullets",
    12: "Weitere-Anmerkungen",
    13: "Achtung",
    14: "Variante",
    15: "Hinweis",
    16: "Typ",
    17: "Abmessungen",
    18: "Leistung-Leuchtmittel"
    }
CURR_PATH = "d:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-06-16\\"
EXPORTS4TRANS = CURR_PATH + "exports4trans\\"
EXPORTED_COLUMNS = CURR_PATH + "4trans_sheets\\"
TRANSCOLUMNS = list(XLS_COLUMNS.keys())


def find_untrans(files4trans_path, columns):
    '''
    Provide the path to the directory
    with xlsx exports for trans.
    Source cells that require translation
    will be exported
    '''
    dir_list = os.listdir(files4trans_path) 
    if os.path.exists(files4trans_path):
        print("Folder exist.")
    else:
        print("Folder does not exist.")

    print(f"dir_list= {dir_list}")
    # get all translated files and build trans memmory as dataframe
    for file in dir_list:
        file4trans_path = files4trans_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        
        for tab in wb.sheetnames: #for each workseet in a workbook
            ws = wb[tab]
            row_limit = ws.max_row
            print(f"Number of rows: {row_limit}")
            print(f"Current sheet: {ws}")
            print(f"Current tab: {tab}")
    
            for column in columns:
                strings4trans = pd.DataFrame(columns=["DE", "PL"])
                curr_row = 2
                while curr_row < row_limit:
                    deu_cell_1 = ws.cell(row=curr_row, column=column)
                    deu_cell_2 = ws.cell(row=curr_row+1, column=column)
                    pol_cell_1 = ws.cell(row=curr_row+2, column=column)
                    pol_cell_2 = ws.cell(row=curr_row+3, column=column)
                                        
                    if deu_cell_1.value != None and pol_cell_1.value is None:
                        new_row = pd.DataFrame({"DE": [deu_cell_1.value], 
                                                "PL": [pol_cell_1.value]})
                        strings4trans = pd.concat([strings4trans, new_row], ignore_index=True)
                        # print(f"curr_row= {curr_row}, deu_cell1_value= {deu_cell_1.value}")
                        # print(f"curr_row+2= {curr_row+2}, pol_cell1_value= {pol_cell_1.value}")
                    else:
                        pass

                    if deu_cell_2.value != None and pol_cell_2.value is None:
                        new_row = pd.DataFrame({"DE": [deu_cell_2.value], 
                                                "PL": [pol_cell_2.value]})
                        strings4trans = pd.concat([strings4trans, new_row], ignore_index=True)
                        # print(f"curr_row+1= {curr_row+1}, deu_cell2_value= {deu_cell_2.value}")
                        # print(f"curr_row+3= {curr_row+3}, pol_cell2_value= {pol_cell_2.value}")
                    else:
                        pass

                    curr_row += 4
                # strings4trans = strings4trans[["DE", "PL"]] #select only DE and PL columns
                print(strings4trans.head())
                print(strings4trans.shape)

                # save the dataframe to excel
                if strings4trans.shape[0] == 0:
                    pass
                else:
                    file_path = EXPORTED_COLUMNS + file[:-5] + "-" + tab + "-" + "COL" + "-" + str(XLS_COLUMNS[column]) + ".xlsx"
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        strings4trans.to_excel(writer, sheet_name=str(XLS_COLUMNS[column]), index=False)
                

find_untrans(EXPORTS4TRANS, TRANSCOLUMNS)