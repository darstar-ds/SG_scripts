import pandas as pd
import os
import openpyxl
import datetime

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Langbeschreibung",
    11: "Ergaenzung-Bullets",
    12: "Weitere-Anmerkungen",
    13: "Achtung",
    14: "Variante",
    15: "Hinweis",
    16: "Typ",
    17: "Abmessungen",
    18: "Leistung-Leuchtmittel"
    }

CURR_PATH = "d:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-07-26\\"
EXPORTS4TRANS = CURR_PATH + "exports4trans\\"
EXPORTED_COLUMNS = CURR_PATH + "4trans_sheets\\"
TRANSCOLUMNS = list(XLS_COLUMNS.keys())


def extract_strings(worksheet, column_no, row_limit):
    prev_time = datetime.datetime.now()
    curr_row = 2
    DE_strings = pd.DataFrame(columns=["datenart", "sprache", "produktno", "navisionid", "text"])
    PL_strings = pd.DataFrame(columns=["datenart", "sprache", "produktno", "navisionid", "text"])
    while curr_row < row_limit:
        if worksheet.cell(row=curr_row, column=2).value == "deu" and \
            worksheet.cell(row=curr_row, column=column_no).value is not None:
            source_datenart = worksheet.cell(row=curr_row, column=1).value
            source_sprache = worksheet.cell(row=curr_row, column=2).value
            source_produktno = worksheet.cell(row=curr_row, column=3).value
            source_navisionid = worksheet.cell(row=curr_row, column=4).value
            source_text = worksheet.cell(row=curr_row, column=column_no).value
            
            data = {"datenart": [source_datenart],
                     "sprache": [source_sprache],
                     "produktno": [source_produktno],
                     "navisionid": [source_navisionid],
                     "text": [source_text]}
            
            new_row = pd.DataFrame(data)
            DE_strings = pd.concat([DE_strings, new_row], ignore_index=True)
            
        elif worksheet.cell(row=curr_row, column=2).value == "pol" and \
            worksheet.cell(row=curr_row, column=column_no).value is None:
            target_datenart = worksheet.cell(row=curr_row, column=1).value
            target_sprache = worksheet.cell(row=curr_row, column=2).value
            target_produktno = worksheet.cell(row=curr_row, column=3).value
            target_navisionid = worksheet.cell(row=curr_row, column=4).value
            target_text = worksheet.cell(row=curr_row, column=column_no).value

            
            data = {"datenart": [target_datenart],
                     "sprache": [target_sprache],
                     "produktno": [target_produktno],
                     "navisionid": [target_navisionid],
                     "text": [target_text]}
            
            new_row = pd.DataFrame(data)
            PL_strings = pd.concat([PL_strings, new_row], ignore_index=True)
            
        else:
            pass
        curr_row += 1
        if curr_row % 5000 == 0:
            curr_time = datetime.datetime.now()
            time_delta = curr_time - prev_time
            print(f"Column: {XLS_COLUMNS[column_no]}, Progress: {curr_row}/{row_limit}, Time: {curr_time}, Delta: {time_delta}")
            prev_time = curr_time

        DE_strings_unique = DE_strings.drop_duplicates(keep="first")
        PL_strings_unique = PL_strings.drop_duplicates(keep="first")
    return DE_strings_unique, PL_strings_unique

def select_untranslated(DE_strings, PL_strings):
    strings_no_PL = pd.DataFrame(columns=["datenart", "sprache", "produktno", "navisionid", "text"])
    for index, row in DE_strings.iterrows():
        # print(f"full_row = {row}")
        # print(f"row0 = {row['datenart']}")
        # print(f"Is row0 in PL_strings? {row['datenart'] in PL_strings['datenart'].values and row['produktno'] in PL_strings['produktno'].values and row['navisionid'] in PL_strings['navisionid'].values}")
        # print(f"row2 = {row['produktno']}")
        # print(f"row3 = {row['navisionid']}")
        if row["datenart"] in PL_strings["datenart"].values and \
           row["produktno"] in PL_strings["produktno"].values and \
           row["navisionid"] in PL_strings["navisionid"].values:
            # print(type(row))
            DE_string = pd.DataFrame(row).transpose()
            # print(DE_string.shape)
            # print(DE_string)
            # DE_string.columns = ["datenart", "sprache", "produktno", "navisionid", "text"]
            # print(type(DE_string))
            # print(DE_string)
            strings_no_PL = pd.concat([strings_no_PL, DE_string], ignore_index=True, axis = 0)
            # print(strings_no_PL)
    return strings_no_PL

def select_strings4trans(strings_no_PL):
    strings4trans = pd.DataFrame(columns=['DE', 'PL'])
    strings4trans["DE"] = strings_no_PL["text"]
    # print(strings4trans.head())
    strings4trans["PL"] = ""
    strings4trans_unique = strings4trans.drop_duplicates(keep="first")
    return strings4trans_unique

def find_untrans(files4trans_path, columns):
    '''
    Provide the path to the directory
    with xlsx exports for trans.
    Source cells that require translation
    will be exported
    '''

    for file in dir_list: # get all untranslated entries and build separate workbook
        file4trans_path = files4trans_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        
        for tab in wb.sheetnames: #for each worksheet in a workbook
            ws = wb[tab]
            row_limit = ws.max_row
            print(f"Number of rows: {row_limit} \nCurrent sheet: {ws} \nCurrent tab: {tab}")
            
            # verify the names/number of columns in the spreadsheet
            list_with_values=[]
            for cell in ws[1]:
                list_with_values.append(cell.value)
            print(f"Columns in {tab} tab: {list_with_values}")
            print(f"Columns no in {tab} tab: {len(list_with_values)}")
    
            for transcolumn in columns:
                DE_strings, PL_strings = extract_strings(ws, transcolumn, row_limit)
                # print(DE_strings.head())
                # print(DE_strings.shape)
                # print(PL_strings.head())
                # print(PL_strings.shape)
                strings_no_PL = select_untranslated(DE_strings, PL_strings)
                # print(strings_no_PL.head())
                # print(strings_no_PL.shape)
                strings4trans = select_strings4trans(strings_no_PL)
                print(f"Tab: {tab}, Column: {XLS_COLUMNS[transcolumn]}, unique dataframe head: {strings4trans.head()}")
                print(f"Unique dataframe shape: {strings4trans.shape}")
  
                # save the dataframe to excel
                if strings4trans.shape[0] == 0:
                    pass
                else:
                    file_path = EXPORTED_COLUMNS + file[:-5] + "-" + tab + "-" + "COL" + "-" + str(XLS_COLUMNS[transcolumn]) + "_unique.xlsx"
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        strings4trans.to_excel(writer, sheet_name=str(XLS_COLUMNS[transcolumn]), index=False)
                

dir_list = os.listdir(EXPORTS4TRANS) 
if os.path.exists(EXPORTS4TRANS):
    print(f"Folder {EXPORTS4TRANS} exist.")
    print(f"dir_list= {dir_list}")
    find_untrans(EXPORTS4TRANS, TRANSCOLUMNS)
else:
    print(f"Folder {EXPORTS4TRANS} does not exist.")

