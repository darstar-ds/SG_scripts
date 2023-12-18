import pandas as pd
import os
import csv
import openpyxl
import datetime

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Einleitung",
    11: "Langbeschreibung",
    12: "Ergaenzung-Bullets",
    13: "Weitere-Anmerkungen",
    14: "Weitere-Besonderheiten",
    15: "Achtung",
    # 16: "Variante",
    16: "Hinweis",
    17: "Typ",
    18: "Abmessungen",
    # 20: "Leistung-Leuchtmittel"
    }

CURR_PATH = "e:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-12-01\\"
TRANSLATED_FILES_PATH = CURR_PATH + "translated\\"
FILES4TRANS_PATH = CURR_PATH + "exports4trans\\"

def translated_gloss(translated_files_path):
    '''
    Provide the path to the directory
    with translated xlsx files.
    Empty dataframe "translated_gloss" is created and
    each file is added to the dataframe.
    '''
    
    dir_list = os.listdir(translated_files_path)
    translated_gloss = pd.DataFrame(columns=["DE", "PL"])

    # get all translated files and build trans memory as dataframe
    for file in dir_list:
        trans_file_path = translated_files_path + file
        one_trans_file = pd.ExcelFile(trans_file_path)
        df_one_trans_file = one_trans_file.parse() # convert into dataframe
        df_one_trans_file = df_one_trans_file[["DE", "PL"]] #select only DE and PL columns
        translated_gloss = pd.concat([translated_gloss, df_one_trans_file], axis=0) #add the current file to the gloss
        translated_gloss.to_excel(CURR_PATH + "slownik.xlsx", index=False)
    return translated_gloss

def translate_workbooks(workbooks_path, columns):
    '''
    Provide the path to a folder with xlsx files to be translated.
    Select the columns to be translated.
    '''

    dir_list = os.listdir(workbooks_path)
    missing_keys_csv = CURR_PATH + 'missing_keys.csv'

    prev_time = datetime.datetime.now()
    curr_row = 2

    for file in dir_list:
        file4trans_path = workbooks_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        for tab in wb.sheetnames: #for each worksheet in a workbook
            worksheet = wb[tab]
            row_limit = worksheet.max_row
            print(f"Current tab: {tab}, Number of rows: {row_limit}")
            
            for column_no in columns.keys(): #column is a number
                curr_row = 2
                while curr_row < row_limit:
                    if worksheet.cell(row=curr_row, column=2).value == "deu" and \
                        worksheet.cell(row=curr_row, column=column_no).value is not None:
                        source_datenart = worksheet.cell(row=curr_row, column=1).value
                        # source_sprache = worksheet.cell(row=curr_row, column=2).value
                        source_produktno = worksheet.cell(row=curr_row, column=3).value
                        source_navisionid = worksheet.cell(row=curr_row, column=4).value
                        source_text = worksheet.cell(row=curr_row, column=column_no).value
                        
                        subrow = curr_row + 1

                        while subrow - curr_row < 5:
                            if worksheet.cell(row=subrow, column=2).value == "pol" and \
                                worksheet.cell(row=subrow, column=1).value == source_datenart and \
                                worksheet.cell(row=subrow, column=3).value == source_produktno and \
                                worksheet.cell(row=subrow, column=4).value == source_navisionid and \
                                worksheet.cell(row=subrow, column=column_no).value is None:
                                
                                try:
                                    target_text = transmem_dict[source_text]
                                    worksheet.cell(row=subrow, column=column_no).value = target_text
                                except KeyError:
                                    # Open the CSV file and append the missing key
                                    with open(missing_keys_csv, mode='a', newline='') as csvfile:
                                        csv_writer = csv.writer(csvfile)
                                        csv_writer.writerow([source_text])  # Write the missing key
                                    # Consider printing a message to the console or logging
                                    print(f"Missing key '{source_text}' added to {missing_keys_csv}")

                            
                            subrow += 1

                    curr_row += 1
                    if curr_row % 5000 == 0:
                        curr_time = datetime.datetime.now()
                        time_delta = curr_time - prev_time
                        print(f"Column: {XLS_COLUMNS[column_no]}, Progress: {curr_row}/{row_limit}, Time: {curr_time}, Delta: {time_delta}")
                        prev_time = curr_time

        # remove tables, i.e. remove autofilter
        for item in worksheet.tables.items():
            del worksheet.tables[item[0]]
        wb.save(file4trans_path)


transmem_dict = dict(translated_gloss(TRANSLATED_FILES_PATH).values)

dir_list = os.listdir(FILES4TRANS_PATH) 
if os.path.exists(FILES4TRANS_PATH):
    print(f"Folder exists {FILES4TRANS_PATH}.")
    translate_workbooks(FILES4TRANS_PATH, XLS_COLUMNS)
    print(f"dir_list= {dir_list}")
else:
    print(f"Folder {FILES4TRANS_PATH} does not exist.")
