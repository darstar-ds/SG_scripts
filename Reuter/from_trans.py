import pandas as pd
import os
import openpyxl

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Langbeschreibung",
    11: "Ergaenzung-Bullets",
    12: "Weitere-Anmerkungen",
    13: "Achtung",
    14: "Variante",
    15: "Hinweis",
    16: "Typ",
    17: "Abmessungen",
    18: "Leistung-Leuchtmittel"
    }
# XLS_COLUMNS = {
#     6: "Kurzbeschreibung",
#     7: "Optionstext",
#     8: "",
#     9: "Lieferumfang",
#     10: "Langbeschreibung",
#     11: "Ergaenzung-Bullets",
#     12: "Weitere-Anmerkungen",
#     13: "Achtung",
#     14: "Variante",
#     15: "Hinweis",
#     16: "Typ",
#     17: "Abmessungen",
#     18: "Leistung-Leuchtmittel"
#     }
CURR_PATH = "d:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-06-21\\"
TRANSLATED_FILES_PATH = CURR_PATH + "translated\\"
FILES4TRANS_PATH = CURR_PATH + "exports4trans\\"

def translated_gloss(translated_files_path):
    '''
    Provide the path to the directory
    with translated xlsx files.
    Empty dataframe "translated_gloss" is created and
    each file is added to the dataframe.
    '''
    dir_list = os.listdir(translated_files_path) 
    if os.path.exists(translated_files_path):
        print(f"Folder exist {translated_files_path}.")
    else:
        print(f"Folder {translated_files_path} does not exist.")
    print(f"dir_list= {dir_list}")

    translated_gloss = pd.DataFrame(columns=["DE", "PL"])

    # get all translated files and build trans memmory as dataframe
    for file in dir_list:
        trans_file_path = translated_files_path + file
        one_trans_file = pd.ExcelFile(trans_file_path)
        df_one_trans_file = one_trans_file.parse() # convert into dataframe
        df_one_trans_file = df_one_trans_file[["DE", "PL"]] #select only DE and PL columns
        translated_gloss = pd.concat([translated_gloss, df_one_trans_file], axis=0) #add the current file to the gloss
        translated_gloss.to_excel(CURR_PATH + "slownik.xlsx", index=False)
    return translated_gloss

def translate_workbooks(workbooks_path, columns):
    '''
    Provide the path to a folder with xlsx files to be translated.
    Select the columns to be translated.
    '''

    dir_list = os.listdir(workbooks_path) 
    if os.path.exists(workbooks_path):
        print(f"Folder exist '{workbooks_path}'.")
    else:
        print(f"Folder '{workbooks_path}' does not exist.")
    print(f"dir_list= {dir_list}")

    for file in dir_list:
        file4trans_path = workbooks_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        for tab in wb.sheetnames: #for each worksheet in a workbook
            ws = wb[tab]
            row_limit = ws.max_row
            print(f"Current tab: {tab}, Number of rows: {row_limit}")
            
            for column in columns: #column is a number
                curr_row = 2
                while curr_row < row_limit:
                    deu_cell_1 = ws.cell(row=curr_row, column=column)
                    deu_cell_2 = ws.cell(row=curr_row+1, column=column)
                    pol_cell_1 = ws.cell(row=curr_row+2, column=column)
                    pol_cell_2 = ws.cell(row=curr_row+3, column=column)
                    
                    # if deu_cell_1.value is None:
                    #     pol_cell_1.value = None
                    # else:
                    #     pol_cell_1.value = transmem_dict[deu_cell_1.value]
                    
                    # if deu_cell_2.value is None:
                    #     pol_cell_2.value = None
                    # else:
                    #     pol_cell_2.value = transmem_dict[deu_cell_2.value]
                    
                    if deu_cell_1.value is not None and pol_cell_1.value is None:
                        pol_cell_1.value = transmem_dict[deu_cell_1.value]
                    
                    if deu_cell_2.value is not None and pol_cell_2.value is None:
                        pol_cell_2.value = transmem_dict[deu_cell_2.value]

                    curr_row += 4
            
        # remove tables, i.e. remove autofilter
        for item in ws.tables.items():
            del ws.tables[item[0]]
        wb.save(file4trans_path)





# transmem = translated_gloss(TRANSLATED_FILES_PATH)

transmem_dict = dict(translated_gloss(TRANSLATED_FILES_PATH).values)

translate_workbooks(FILES4TRANS_PATH, XLS_COLUMNS)