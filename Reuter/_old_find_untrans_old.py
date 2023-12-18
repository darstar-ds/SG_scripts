import pandas as pd
import os
import openpyxl

XLS_COLUMNS = {
    6: "Kurzbeschreibung",
    7: "Optionstext",
    9: "Lieferumfang",
    10: "Einleitung",
    11: "Langbeschreibung",
    12: "Ergaenzung-Bullets",
    13: "Weitere-Anmerkungen",
    14: "Weitere-Besonderheiten",
    15: "Achtung",
    16: "Variante",
    17: "Hinweis",
    18: "Typ",
    19: "Abmessungen",
    20: "Leistung-Leuchtmittel"
    }

CURR_PATH = "e:\\Moje dokumenty\\SG_scripts_data\\Reuter\\2023-10-18\\"
EXPORTS4TRANS = CURR_PATH + "exports4trans\\"
EXPORTED_COLUMNS = CURR_PATH + "4trans_sheets\\"
TRANSCOLUMNS = list(XLS_COLUMNS.keys())


def find_untrans(files4trans_path, columns):
    '''
    Provide the path to the directory
    with xlsx exports for trans.
    Source cells that require translation
    will be exported
    '''
    dir_list = os.listdir(files4trans_path) 
    if os.path.exists(files4trans_path):
        print(f"Folder {files4trans_path} exist.")
    else:
        print(f"Folder {files4trans_path} does not exist.")
    print(f"dir_list= {dir_list}")
    
    for file in dir_list: # get all untranslated entries and build separate workbook
        file4trans_path = files4trans_path + file
        wb = openpyxl.load_workbook(file4trans_path)
        
        for tab in wb.sheetnames: #for each worksheet in a workbook
            ws = wb[tab]
            row_limit = ws.max_row
            print(f"Number of rows: {row_limit}")
            print(f"Current sheet: {ws}")
            print(f"Current tab: {tab}")

            # verify the names/number of columns in the spreadsheet
            list_with_values=[]
            for cell in ws[1]:
                list_with_values.append(cell.value)
            print(f"Columns in {tab} tab: {list_with_values}")
            print(f"Columns no in {tab} tab: {len(list_with_values)}")
    
            for column in columns:
                strings4trans = pd.DataFrame(columns=["DE", "PL"])
                curr_row = 2
                while curr_row < row_limit:
                    deu_cell_1 = ws.cell(row=curr_row, column=column)
                    deu_cell_2 = ws.cell(row=curr_row+1, column=column)
                    pol_cell_1 = ws.cell(row=curr_row+2, column=column)
                    pol_cell_2 = ws.cell(row=curr_row+3, column=column)
                                        
                    if deu_cell_1.value != None and pol_cell_1.value is None:
                        new_row = pd.DataFrame({"DE": [deu_cell_1.value], 
                                                "PL": [pol_cell_1.value]})
                        strings4trans = pd.concat([strings4trans, new_row], ignore_index=True)
                    else:
                        pass

                    if deu_cell_2.value != None and pol_cell_2.value is None:
                        new_row = pd.DataFrame({"DE": [deu_cell_2.value], 
                                                "PL": [pol_cell_2.value]})
                        strings4trans = pd.concat([strings4trans, new_row], ignore_index=True)
                    else:
                        pass

                    curr_row += 4
                
                print(f"Tab: {tab}, Column: {XLS_COLUMNS[column]}, initial dataframe head: {strings4trans.head()}")
                print(f"Initial dataframe shape: {strings4trans.shape}")

                # extract unique entries only
                strings4trans_unique = strings4trans.drop_duplicates(keep="first")
                print(f"Tab: {tab}, Column: {XLS_COLUMNS[column]}, unique dataframe head: {strings4trans_unique.head()}")
                print(f"Unique dataframe shape: {strings4trans_unique.shape}")

                # save the dataframe to excel
                if strings4trans_unique.shape[0] == 0:
                    pass
                else:
                    file_path = EXPORTED_COLUMNS + file[:-5] + "-" + tab + "-" + "COL" + "-" + str(XLS_COLUMNS[column]) + "_unique.xlsx"
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        strings4trans_unique.to_excel(writer, sheet_name=str(XLS_COLUMNS[column]), index=False)
                

find_untrans(EXPORTS4TRANS, TRANSCOLUMNS)