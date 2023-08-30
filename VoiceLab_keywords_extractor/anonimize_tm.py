import pandas as pd
import openpyxl


CURR_PATH = "d:\\Moje dokumenty\\SG_scripts_data\\SG\\"
END_CUSTOMERS = "End_Customers.txt"
TM4ANONYM = "SG_legal_EN-PL-738457.xlsx"

def build_end_customers():
    '''
    The function takes the txt file and builds a list of end customers.
    Returns a list.
    '''
    with open(CURR_PATH+END_CUSTOMERS, "r", encoding='utf-8') as file:
        end_customers = file.read().splitlines()
    print(end_customers)
    return end_customers

def build_tm():
    '''
    The function takes XLS file with exportemTM and convert it to dataframe.
    Returns a dataframe.
    '''
    wb = openpyxl.load_workbook(CURR_PATH+TM4ANONYM)
    sh = wb.active
    tm_dataframe = pd.DataFrame(columns=["Source", "Target"])
    print("Converting TM from XLS to DATAFRAME...")
    for i in range(2, sh.max_row+1):
        source = sh.cell(row=i, column=2).value
        target = sh.cell(row=i, column=6).value
        data = {"Source": [source],
                "Target": [target]}
        new_row = pd.DataFrame(data)
        tm_dataframe = pd.concat([tm_dataframe, new_row], ignore_index=True)
        # print(tm_dataframe)
        if i % 5000 == 0:
            print(f"Row: {i}")
    return tm_dataframe

def ec2dict(end_customers):
    '''
    The function takes the list of end customers and convert it to dictionary.
    Returns a dictionary used as a counter.
    '''
    ec_counter = {}
    for customer in end_customers:
        ec_counter[customer] = 0
    print(ec_counter)
    return ec_counter

def anonimize_tm(end_customers, tm, ec_counter):
    '''
    The function takes end customers list and translation memory as dataframe.
    Replaces end customer name with ACME.
    Returns anonimized tm as dataframe.
    '''
    print("Anonimizing the TM...")
    tm_changes = pd.DataFrame(columns=["Source_old", "Source_new", "End Customer"])
    for index, row in tm.iterrows():
        for end_customer in end_customers:
            if end_customer in row["Source"]:
                source_old = row["Source"]
                row["Source"] = row["Source"].replace(end_customer, "ACME")
                source_new = row["Source"]
                tm_changes_line = pd.DataFrame(data=[[source_old, source_new, end_customer]], columns=["Source_old", "Source_new", "End Customer"])
                tm_changes = pd.concat([tm_changes, tm_changes_line], ignore_index=True)
                ec_counter[end_customer] = ec_counter[end_customer] + 1
        if index % 5000 == 0:
            print(f"Row: {index}")
    for i in ec_counter:
        if ec_counter[i] != 0:
            print(i, ec_counter[i])
    file_path = CURR_PATH + "tm_changes.xlsx"
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        tm_changes.to_excel(writer, sheet_name="tm_changes", index=False)
    return tm, tm_changes

end_customers = build_end_customers()
ec_counter = ec2dict(end_customers)
tm_dataframe = build_tm()
anonimized_tm, tm_changes = anonimize_tm(end_customers, tm_dataframe, ec_counter)
print(tm_changes.head(100))